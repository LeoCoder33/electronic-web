import os.path
import urllib
from datetime import datetime
from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl.workbook import Workbook

from algorithm.util import plan_evaluation
from flask import Flask, request, g, make_response

app = Flask(__name__)

app.config['excel_data1'] = None
app.config['excel_data2'] = None


@app.route('/evaluate', methods=["POST"])
def evaluate():
    files = request.files
    # 网架数据
    architecture = files.get('model')
    # 已有容量
    existPV = files.get('PV')
    # 报装容量
    PackingPV = files.get('PVmethod')
    # 新负荷数据
    loadData = files.get('load')
    # 光伏数据
    PVdata = files.get('PVdata')

    evaluate_result, cross_result = plan_evaluation(architecture, existPV, PackingPV, loadData, PVdata)

    # 把evaluation_data封装成xslx文件
    # 将数组写入Excel文件
    # Create a new Excel workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active
    # Write the NumPy array data to the Excel sheet
    for row_idx, row_data in enumerate(evaluate_result):
        for col_idx, value in enumerate(row_data):
            sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)
    # Create an in-memory buffer to hold the Excel file
    buffer = BytesIO()
    # Save the workbook to the buffer
    workbook.save(buffer)
    # Store the buffer in the g object for future requests
    app.config['excel_data1'] = buffer.getvalue()
    capacity = {
        'node': ['节点'] + list(evaluate_result[:, 0]),
        'loadPV': ['报装容量'] + list(evaluate_result[:, 2]),
        'existPV': ['已有容量'] + list(evaluate_result[:, 1])
    }

    capacity_table = []
    for i in range(1, len(capacity['node'])):
        capacity_table.append({
            'node': capacity['node'][i],
            'loadPV': capacity['loadPV'][i],
            'existPV': capacity['existPV'][i]
        })

    return {
        'capacity': capacity,
        'capacity_table': capacity_table,
        'cross_result': cross_result
    }


@app.route('/downloadEvalFile', methods=["GET"])
def downloadEvalFile():
    # Retrieve the in-memory buffer from the g object
    buffer = app.config['excel_data1']
    # Create a bytes string for the response
    rsp = make_response(buffer)
    # 获取当前日期时间
    todayDate = datetime.now().strftime("%y-%m-%d-%H-%M")
    filename = todayDate + "光伏优化方案.xlsx"
    # Set the right header and mimetype for the response
    rsp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    rsp.headers['Content-Disposition'] = 'attachment; filename*=UTF-8\'\'{}'.format(urllib.parse.quote(filename))
    return rsp


@app.route('/calculate', methods=["POST"])
def calculate():
    # 判断传来的参数’method‘为’台区改建‘还是’台区扩建‘
    method = request.form.get('method')
    if method == '台区改建':
        evaluation_data = np.array([
            [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18],
            [6.17, 8.00, 6.00, 4.00, 9.33, 9.00, 0.63, 0.00, 2.87, 9.00, 0.00, 0.00, 0.65, 1.00, 8.00, 8.00, 3.00, 7.00]
        ])
        capacity = {
            'node': ['节点'] + [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18],
            'optimizePV': ['优化容量'] + [6.17, 8.00, 6.00, 4.00, 9.33, 9.00, 0.63, 0.00, 2.87, 9.00, 0.00, 0.00, 0.65, 1.00, 8.00, 8.00, 3.00, 7.00]
        }
    else:
        evaluation_data = np.array([
            [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18],
            [8.00, 7.00, 4.89, 4.00, 2.69, 5.00, 3.61, 0.94, 0.97, 7.25, 10.00, 1.33, 0.00, 1.00, 1.00, 8.69, 7.60, 7.03]
        ])
        capacity = {
            'node': ['节点'] + [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18],
            'optimizePV': ['优化容量'] + [8.00, 7.00, 4.89, 4.00, 2.69, 5.00, 3.61, 0.94, 0.97, 7.25, 10.00, 1.33, 0.00, 1.00, 1.00, 8.69, 7.60, 7.03]
        }

    # 把calculation_data封装成xslx文件
    # 将数组写入Excel文件
    # Create a new Excel workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active
    # Write the NumPy array data to the Excel sheet
    for row_idx, row_data in enumerate(evaluation_data):
        for col_idx, value in enumerate(row_data):
            sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)
    # Create an in-memory buffer to hold the Excel file
    buffer = BytesIO()
    # Save the workbook to the buffer
    workbook.save(buffer)
    # Store the buffer in the g object for future requests
    app.config['excel_data2'] = buffer.getvalue()

    capacity_table = []
    for i in range(1, len(capacity['node'])):
        capacity_table.append({
            'node': capacity['node'][i],
            'optimizePV': capacity['optimizePV'][i]
        })
    cross_result = -1
    return {
        'capacity': capacity,
        'capacity_table': capacity_table,
        'cross_result': cross_result
    }


# download excel
@app.route('/downloadCalFile', methods=["GET"])
def downloadCalFile():
    # Retrieve the in-memory buffer from the g object
    buffer = app.config['excel_data2']
    # Create a bytes string for the response
    rsp = make_response(buffer)
    # 获取当前日期时间
    todayDate = datetime.now().strftime("%y-%m-%d-%H-%M")
    filename = todayDate + "光伏优化方案.xlsx"
    # Set the right header and mimetype for the response
    rsp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    rsp.headers['Content-Disposition'] = 'attachment; filename*=UTF-8\'\'{}'.format(urllib.parse.quote(filename))
    return rsp

# 请求返回迭代数据
@app.route('/iteration', methods=["GET"])
def iteration():
    # 读取request中的参数’iter_num‘，’method‘
    iter_num = int(request.args.get('iter_num'))
    method = request.args.get('method')
    # 根据method判断是台区改建还是台区扩建
    if method == '台区改建':
        # np读取’RebuildDataProcess.xlsx‘文件中的数据
        data = np.array(pd.read_excel('RebuildDataProcess.xlsx', header=None).astype(float))
    else:
        # np读取’ExpansionDataProcess.xlsx‘文件中的数据
        data = np.array(pd.read_excel('ExpansionDataProcess.xlsx', header=None).astype(float))
    # 取出第iter_num列数据并转换成列表
    data = data[:, iter_num].tolist()
    loss = ['loss'] + [i*1e6 for i in data[0:500]]
    cost = ['cost'] + [i*1e4 for i in data[500:1000]]
    fluctuation = ['fluctuation'] + data[1000:1500]
    data_source = [loss, fluctuation, cost]
    return {
        'data_source': data_source,
        'iter_num': iter_num+1
    }



if __name__ == '__main__':
    app.run()
